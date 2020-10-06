#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time     :2020-10-04 12:16
# @Author   :wangqinghua
# @File     : du_excel.py
# @Software : PyCharm


#  操作读取excel接口用例 再把结果写回excel
from openpyxl import load_workbook

class Doexcel:
    def get_data(self,file_name,sheet_name):

        wb = load_workbook(file_name)
        sheet = wb[sheet_name]

        test_data=[]
        for i in range(2,sheet.max_row+1):
            row_data={}
            row_data['case_id'] = sheet.cell(i, 1).value
            row_data['url']=sheet.cell(i, 2).value
            row_data['data'] = sheet.cell(i,3).value
            row_data['title'] = sheet.cell(i, 4).value
            row_data['http_method'] = sheet.cell(i, 5).value
            test_data.append(row_data)

        return test_data
    def write_back(self,file_name,sheet_name,i,value):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i,6).value=value
        wb.save(file_name) #保存结果

#第一种 读取登录表单
# if __name__ == '__main__':   # 读取文件的3种方法
# #     test_date = Doexcel().get_data(r"E:\tmp\jichuke\Actual_practice\case_one\test.xlsx","")  # 在当前文件夹下创建 默认取 excel文件名就可以
# #     # test_date = Doexcel().get_data(r"E:\tmp\jichuke\Actual_practice\test\te.xlsx", "login")  # 切换文件夹后使用绝对路径
# #     # test_date = Doexcel().get_data("../test/te.xlsx","login") # 相对路径读取文件
# #     print(test_date)



# 第二种 读取创建课程表单
if __name__ == '__main__':   # 读取文件的3种方法
    test_date = Doexcel().get_data("test.xlsx","add")  # 在当前文件夹下创建 默认取 excel文件名就可以
    # test_date = Doexcel().get_data(r"E:\tmp\jichuke\Actual_practice\case_one\test.xlsx", "login")  # 切换文件夹后使用绝对路径
    # test_date = Doexcel().get_data("../test/te.xlsx","login") # 相对路径读取文件
    print(test_date)